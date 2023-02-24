# mearge 관련 오류가 나올시 누수감면 시트에 병합된 셀이 있는 확인해주세요
# 연번 순서대로 입력하는 것이 아닌 누수감면 조정내역의 시트 순서대로 입력된다.

from openpyxl import load_workbook 
from openpyxl import Workbook


#직접 입력 작업을 수행하는 메소드
def write_data_to_excel(num,k,ws):
    global input_cnt

    find_row=0

    for j in range(6,new_ws_max_row):
        if new_ws.cell(j,1).value == num:
            find_row = j
            break
            
    #입력
    if find_row != 0:

        #누수감면 파일에 입력값이 이미 있는 연번은 입력 작업을 실행하지 않는다. (고객번호 기준으로 판단함)
        if new_ws.cell(find_row,4).value == None:

            #주소
            new_ws["F"+str(find_row)] = ws.cell(k+1,2 ).value
            
            #성명
            new_ws["G"+str(find_row)] = ws.cell( k+2,2).value
            
            #고객번호
            new_ws["D"+str(find_row)] = str(ws.cell(k+2,8).value)[-6:]
            

            #상수도 - 업종, 당초금액(량, 금액), 정당금액(량, 금액)
            new_ws["I"+str(find_row)] = ws.cell(k+6,2 ).value
            new_ws["J"+str(find_row)] = ws.cell(k+6,3).value
            new_ws["K"+str(find_row)] = ws.cell(k+6,4).value
            new_ws["L"+str(find_row)] = ws.cell(k+6,5).value
            new_ws["M"+str(find_row)] = ws.cell(k+6,6).value

            
            #하수도 - 당초금액(량, 금액), 정당금액(량, 금액)
            new_ws["P"+str(find_row)] = ws.cell(k+7,3).value
            new_ws["Q"+str(find_row)] = ws.cell(k+7,4).value
            new_ws["R"+str(find_row)] = ws.cell(k+7,5).value
            new_ws["S"+str(find_row)] = ws.cell(k+7,6).value
    

            #물 이용금 - 당초금액(량, 금액), 정당금액(량, 금액)
            new_ws["V"+str(find_row)] = ws.cell(k+8,3 ).value
            new_ws["W"+str(find_row)] = ws.cell(k+8,4).value
            new_ws["X"+str(find_row)] = ws.cell(k+8,5).value
            new_ws["Y"+str(find_row)] = ws.cell(k+8,6).value
            
            input_cnt += 1
        
        else:
            print("해당 행은 이미 입력된 값이 있습니다 -> 연번 : "+str(num)+" 고객번호 : "+str(new_ws.cell(find_row,4).value)+"\n")


    else:
        #오류가 발생시 저장되지 않도록한다.
        bool_test=True
        print("오류",num)




#해당 누수감면 조정내역의 시트를 받아 해당 연번을 write_data_to_excel로 넘겨주는 메소드
def fill_excel(sheets_name):
    print("\n시트: "+sheets_name+" 입력 중\n")
    ws = wb[sheets_name]
    for i in range(1,ws.max_row):
        ws_cell_i_value= ws.cell(i,1).value
        if  isinstance(ws_cell_i_value,int) and ws.cell(i+11,1).value == None:

            if ws_cell_i_value > max_num:
                print(ws_cell_i_value, "작업 실행")
                write_data_to_excel(ws_cell_i_value,i,ws)
                i+=11




if __name__ == '__main__':

    #파일 경로 입력
    filename1=r'C:\excel_auto\discount\누수감면 조정내역(2023) - 복사본.xlsx'
    filename2=r'C:\excel_auto\discount\누수감면(2023년)_test.xlsx'


    wb =load_workbook(filename1)
    new_wb=load_workbook(filename2)

    wb_sheets=wb.sheetnames

    new_wb_sheets=new_wb.sheetnames

    bool_test= False

    #입력할 누수감면 시트 명
    sheet_name='2월'
    new_ws=new_wb[sheet_name]
    input_cnt = 0

    new_ws_max_row =376

    # 누수감면에 입력되어 있는 마지막 연번 ( *max_num의 값 보다 높은 연번만 값이 입력됨* )
    max_num = 141

    for sheet in wb_sheets:
        fill_excel(sheet)

    #파일 저장 
    if not bool_test:
        new_wb.save(filename =filename2)
        wb.save(filename =filename1)
        print("\n최종 입력 수 : "+str(input_cnt))

    wb.close()
    new_wb.close()
