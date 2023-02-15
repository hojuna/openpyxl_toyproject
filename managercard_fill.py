from openpyxl import load_workbook 


wb =load_workbook('c:/excel_auto/테스트1_연동2.xlsx')
ws = wb['관리카드']

k=1
# 400번까지 채우기
for i in range(1,400):
    num_str='$A'+str(k)
    # r관리번호
    ws['B'+str(3+k)]='="관리번호 제"&'+num_str+'&" 호"'
    
    # 수용가명
    ws['D'+str(4+k)]= '/'
    
    # 구경
    ws['G'+str(4+k)]= '=VLOOKUP('+num_str+',미교체내역!$A:$P,12,FALSE)'

    # 고객번호
    ws['D'+str(5+k)]= '=VLOOKUP('+num_str+',미교체내역!$A:$P,4,FALSE)'
    
    # 기물번호
    ws['G'+str(5+k)]= '=VLOOKUP('+num_str+',미교체내역!$A:$P,6,FALSE)'
    
    # 주소
    ws['D'+str(6+k)]= '=VLOOKUP('+num_str+',미교체내역!$A:$P,9,FALSE)'
    
    
    #교체불가사유 
    ws['D'+str(7+k)]= '=VLOOKUP('+num_str+',미교체내역!$A:$P,16,FALSE)'
    k+=14

wb.save(filename ='테스트1_연동2.xlsx')