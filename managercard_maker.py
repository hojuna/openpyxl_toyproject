from openpyxl import load_workbook 
from openpyxl import Workbook
from tkinter.messagebox import * 

# 엑셀업무자동화 v0.1 ( 23.02.08 demo )

filename1=""
filename2=""


# ex
def load_filename():
    global filename1,filename2
    with open('managercard_log.txt','r',encoding="UTF-8") as f:
        filename1=(f.readline().strip())

    with open('managercard_log2.txt','r',encoding="UTF-8") as f:
        filename2=f.readline().strip()


def crad_make():
    cnt=0
    load_filename()
    try:
        wb =load_workbook(filename1)
        new_wb=load_workbook(filename2)

    except:
        showerror("오류","오류가 발생했습니다. \n파일 경로를 확인해주세요")

    wb_sheets=wb.sheetnames
    ws = wb[wb_sheets[0]]
    new_ws_sheets=new_wb.sheetnames
    new_ws=new_wb[new_ws_sheets[0]]


    try:
        for i in range(1,new_ws.max_row):
            if new_ws.cell(i,1).value == None:
                cnt=i
                break  
    except:
        showerror("오류","오류가 발생했습니다. \n필터정렬 해지해주세요")        


    for x in range(1,ws.max_row):
        dd=ws.cell(x,4).value
        if dd!=None:
            if ws.cell(x,5).value=='부':
                new_ws["A"+str(cnt)]=cnt-1
                new_ws["C"+str(cnt)]=ws.cell(x,1).value
                new_ws["D"+str(cnt)]=ws.cell(x,2).value
                new_ws["F"+str(cnt)]=ws.cell(x,6).value
                new_ws["G"+str(cnt)]=ws.cell(x,7).value      
                new_ws["I"+str(cnt)]=ws.cell(x,8).value
                new_ws["J"+str(cnt)]=ws.cell(x,10).value        
                new_ws["L"+str(cnt)]=ws.cell(x,11).value    
                new_ws["N"+str(cnt)]=ws.cell(x,12).value                
                new_ws["P"+str(cnt)]=ws.cell(x,4).value
                ws["E"+str(x)]='가'
                cnt+=1

    try:
        new_wb.save(filename =filename2)
        wb.save(filename =filename1)
    except:
        showerror("오류","오류가 발생했습니다. \n파일을 닫고 실행해주세요")
    new_wb.close()
    wb.close()


import tkinter as tk
from tkinter.filedialog import askopenfilename

root = tk.Tk()
root.title('관리카드 편집기')
root.resizable(False,False)
root.geometry("1000x400+100+100")

load_filename()



#교체예정 파일 경로 찾기
def open_file():
    filename = askopenfilename(filetypes=(("Excel files",".xlsx .xls"), ('All files','*.*')))
    if filename:
        with open('managercard_log.txt','w',encoding="UTF-8") as f:
            f.truncate(0)
            f.write(filename+"\n")
            # print(filename)
        btn_text.set("현재 설정\n\n"+filename)
        
#관리카드 파일 경로 찾기
def open_file2():
    filename = askopenfilename(filetypes=(("Excel files",".xlsx .xls"), ('All files','*.*')))
    if filename:
        with open('managercard_log2.txt','w',encoding="UTF-8") as f:
            f.truncate(0)
            f.write(filename+"\n")
        btn_text2.set("현재 설정\n\n"+filename)

btn_text = tk.StringVar()
btn_text.set("교체~ 현재 경로\n\n"+filename1)

btn_text2 = tk.StringVar()
btn_text2.set("관리~ 현재 경로\n\n"+filename2)


#Label
widget2 =tk.Button(root,text="교체예정계량기 파일 선택",  fg="white",
    command=open_file,
    # textvariable=btn_text,
    bg="#34A2Fe",
    width=40,
    height=5)
widget2.grid(row=1,column=0,padx=20)

#Label
widget3 =tk.Button(root,text="관리카드 파일 선택",  fg="white",
    command=open_file2,
    # textvariable=btn_text2,
    bg="#34A2Fe",
    width=40,
    height=5)
widget3.grid(row=2,column=0,padx=30)

widget4 =tk.Label(
    root,
    # text=btn_text,
    textvariable=btn_text,
    fg="white",
    bg="#34A2Fe",
    width=40,
    height=5
    )
widget4.grid(row=1,column=10,padx=10)

widget5 =tk.Label(
    root,
    # text=btn_text2,
    textvariable=btn_text2,
    fg="white",
    bg="#34A2Fe",
    width=40,
    height=5
    )
widget5.grid(row=1,column=20,padx=10)

#실행버튼
widget6 =tk.Button(root,text="실행",  fg="white",
    command=crad_make,
    # textvariable=btn_text,
    bg="#34A2Fe",
    width=40,
    height=5)
widget6.grid(row=2,column=10,padx=20)

root.mainloop()