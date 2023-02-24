# pyinstaller managercard_maker_sys.py --noconsole --onefile --add-data "C:\excel_auto\managercard_log.txt;." --add-data "C:\excel_auto\managercard_log2.txt;."
from openpyxl import load_workbook
from openpyxl import Workbook
from tkinter.messagebox import *
from tkinter import messagebox
import sys
import os
from datetime import date, datetime
import tkinter as tk
from tkinter.filedialog import askopenfilename


# 엑셀업무자동화 v0.1 ( 23.02.08 demo )     기본 기능
# 엑셀업무자동화 v0.2 ( 23.02.16 demo )     완료,추진사항 입력기능 개발 및 경로 수정

# 기능 정리
# 로그파일을 활용해 최근 열린 파일이 자동으로 잡힘
# 가,부를 기준으로 관리카드 생성
# 완을 기준으로 해당 고객번호가 관리카드에 존재한다면 완료 정보 입력
# 추진사항(날짜)를 입력
# 총 입력된 사항을 마무리 알림으로 출력

filename1 = ""
filename2 = ""
day_type = date.today()

def load_filename():
    global filename1, filename2

    with open('managercard_log.txt','r',encoding="UTF-8") as f:
        filename1=(f.readline().strip())

    with open('managercard_log2.txt','r',encoding="UTF-8") as f:
        filename2=f.readline().strip()


def crad_make():
    new_ws_last_row = 0
    card_cnt = 0
    completion_cnt = 0
    date_cnt = 0
    find_row_number = 0
    load_filename()
    try:
        wb = load_workbook(filename1)
        new_wb = load_workbook(filename2)

    except:
        showerror("오류", "오류가 발생했습니다. \n파일 경로를 확인해주세요")

    wb_sheets = wb.sheetnames
    ws = wb[wb_sheets[0]]
    new_ws_sheets = new_wb.sheetnames
    new_ws = new_wb[new_ws_sheets[0]]

    try:
        for i in range(1, new_ws.max_row):
            if new_ws.cell(i, 1).value == None:
                basics_new_ws_last_row = i
                break
    except:
        showerror("오류", "오류가 발생했습니다. \n필터정렬 해지해주세요")

    # 현재 추진사항 날짜가 관리카드 파일에 이미 있는지 검사
    def find_date(find_row, date):
        for i in range(3):
            new_ws_value_temp = new_ws.cell(find_row, 16+i).value
            if new_ws_value_temp != None:
                if new_ws_value_temp == str(date.strftime('%Y-%m-%d')):
                    return True
        return False

    new_ws_last_row = basics_new_ws_last_row
    for x in range(1, ws.max_row):
        ws_5_value = ws.cell(x, 5).value
        ws_4_value = ws.cell(x, 4).value
        ws_3_value = ws.cell(x, 3).value
        if ws_4_value != None:
            # 관리카드 입력
            if ws_5_value == '부':

                card_cnt += 1
                new_ws["A"+str(new_ws_last_row)] = new_ws_last_row-1
                new_ws["C"+str(new_ws_last_row)] = ws.cell(x, 1).value
                new_ws["D"+str(new_ws_last_row)] = ws.cell(x, 2).value
                new_ws["F"+str(new_ws_last_row)] = ws.cell(x, 6).value
                new_ws["G"+str(new_ws_last_row)] = ws.cell(x, 7).value
                new_ws["I"+str(new_ws_last_row)] = ws.cell(x, 8).value
                new_ws["J"+str(new_ws_last_row)] = ws.cell(x, 10).value
                new_ws["L"+str(new_ws_last_row)] = ws.cell(x, 11).value

                new_ws["O"+str(new_ws_last_row)] = ws_4_value

                if ws_3_value != None and isinstance(ws_3_value, datetime):
                    new_ws["P"+str(new_ws_last_row)
                           ] = ws_3_value.strftime('%Y-%m-%d')
                    date_cnt += 1

                ws["E"+str(x)] = '가'
                new_ws_last_row += 1

            # 완료 사항 입력
            elif ws_5_value == '완':
                customerNumber = ws.cell(x, 2).value
                find_bool = False
                completion_empty = False
                for i in range(1, basics_new_ws_last_row):
                    if customerNumber == new_ws.cell(i, 4).value:
                        if new_ws.cell(i, 19).value != None:
                            completion_empty = True
                        find_bool = True
                        find_row_number = i
                if find_bool:
                    if not completion_empty:
                        completion_cnt += 1
                        new_ws["S"+str(find_row_number)] = ws.cell(x, 4).value
                    find_bool = False
                else:
                    showerror("오류", "완료 표시된 고객번호 : *"+str(customerNumber) +
                              "*이 관리카드에 존재하지 않습니다. \n 해당 고객번호를 확인해주세요")
        
        if ws_3_value != None:
            if isinstance(ws_3_value, datetime) and ws_5_value == '가':
                customerNumber = ws.cell(x, 2).value
                # print(customerNumber)
                column_of_full = True

                for i in range(1, basics_new_ws_last_row):
                    if customerNumber == new_ws.cell(i, 4).value:
                        find_row_number = i
                
                if find_row_number!=0 and  not find_date(find_row_number, ws_3_value):
                    ws_3_value = ws_3_value.strftime('%Y-%m-%d')
                    for i in range(3):
                        if new_ws.cell(find_row_number, 16+i).value == None:
                            # 비어 있으면 해당 위치에 값을 넣고 column_of_full = false로 바꾸고 break
                            date_cnt += 1
                            if (16 == 16+i):
                                new_ws["P"+str(find_row_number)] = ws_3_value
                                column_of_full = False
                                break
                            elif (17 == 16+i):
                                new_ws["Q"+str(find_row_number)] = ws_3_value
                                column_of_full = False
                                break
                            elif (18 == 16+i):
                                new_ws["R"+str(find_row_number)] = ws_3_value
                                column_of_full = False
                                break

                    if column_of_full:
                        # 추진사항이 모두 가득 차있는 상태라면 한열 땡김
                        date_cnt += 1
                        new_ws["P"+str(find_row_number)
                               ] = new_ws.cell(find_row_number, 17).value
                        new_ws["Q"+str(find_row_number)
                               ] = new_ws.cell(find_row_number, 18).value
                        new_ws["R"+str(find_row_number)] = ws_3_value

                    else:
                        column_of_full = True

    try:
        new_wb.save(filename=filename2)
        wb.save(filename=filename1)
    except:
        showerror("오류", "오류가 발생했습니다. \n파일을 닫고 실행해주세요")

    if completion_cnt != 0:
        if date_cnt != 0:
            showinfo(
                "알림", f"{card_cnt}개의 행이 관리카드로 입력됨 \n{completion_cnt}개의 행의 완료사항이 입력됨 \n{date_cnt}개의 행의 추진사항이 입력됨")
        else:
            showinfo(
                "알림", f"{card_cnt}개의 행이 관리카드로 입력됨 \n{completion_cnt}개의 행의 완료사항이 입력됨")
    elif completion_cnt == 0:
        if date_cnt != 0:
            showinfo(
                "알림", f"{card_cnt}개의 행이 관리카드로 입력됨 \n{date_cnt}개의 행의 추진사항이 입력됨")
        elif card_cnt != 0:
            showinfo("알림", f"{card_cnt}개의 행이 관리카드로 입력됨")
        else:
            showinfo("알림", f"아무것도 입력되지 않음")

    new_wb.close()
    wb.close()


root = tk.Tk()
root.title('관리카드 편집기')
root.resizable(False, False)
root.geometry("1000x400+100+100")



load_filename()


# 교체예정 파일 경로 찾기
def open_file():
    filename = askopenfilename(filetypes=(
        ("Excel files", ".xlsx .xls"), ('All files', '*.*')))

    if filename:
        with open('managercard_log.txt','w',encoding="UTF-8") as f:
            f.truncate(0)
            f.write(filename+"\n")

        btn_text.set("현재 설정\n\n"+filename)

# 관리카드 파일 경로 찾기
def open_file2():

    filename = askopenfilename(filetypes=(("Excel files",".xlsx .xls"), ('All files','*.*')))
    if filename:
        with open('managercard_log2.txt','w',encoding="UTF-8") as f:
            f.truncate(0)
            f.write(filename+"\n")
        btn_text2.set("현재 설정\n\n"+filename)

btn_text = tk.StringVar()
btn_text.set("교체~ 현재 경로\n\n"+filename1)

btn_text2 = tk.StringVar()
btn_text2.set("관리~ 현재 경로\n\n"+filename2)


# Label
widget2 = tk.Button(root, text="교체예정계량기 파일 선택",  fg="white",
                    command=open_file,
                    # textvariable=btn_text,
                    bg="#34A2Fe",
                    width=40,
                    height=5)
widget2.grid(row=1, column=0, padx=20)

# Label
widget3 = tk.Button(root, text="관리카드 파일 선택",  fg="white",
                    command=open_file2,
                    # textvariable=btn_text2,
                    bg="#34A2Fe",
                    width=40,
                    height=5)
widget3.grid(row=2, column=0, padx=30)

widget4 = tk.Label(
    root,
    # text=btn_text,
    textvariable=btn_text,
    fg="white",
    bg="#34A2Fe",
    width=40,
    height=5
)
widget4.grid(row=1, column=10, padx=10)

widget5 = tk.Label(
    root,
    # text=btn_text2,
    textvariable=btn_text2,
    fg="white",
    bg="#34A2Fe",
    width=40,
    height=5
)
widget5.grid(row=1, column=20, padx=10)

# 실행버튼
widget6 = tk.Button(root, text="실행",  fg="white",
                    command=crad_make,
                    # textvariable=btn_text,
                    bg="#34A2Fe",
                    width=40,
                    height=5)
widget6.grid(row=2, column=10, padx=20)

root.mainloop()
